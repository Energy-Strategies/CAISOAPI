'''
This is a GUI program desined to pull CAISO OASIS LMP Pricing data,
cleaned and organized into an excel spreadsheet.
'''

import math
import sys
import threading
import time
from datetime import datetime, timedelta
from io import BytesIO
from zipfile import ZipFile
import tkinter as tk
from tkinter import filedialog

import pandas as pd
import requests
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend — safe to use from worker threads
import matplotlib.pyplot as plt
import seaborn as sns

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

from pydantic import ValidationError

from tkcalendar import Calendar
from customtkinter import (CTk, CTkButton, CTkLabel, CTkComboBox, CTkEntry,
                           CTkCheckBox, set_appearance_mode)

# =============================================================================
# CONSTANTS AND CONFIGURATION
# =============================================================================
MARKET_CONFIG = {
    'DAM': {
        'queryname': 'PRC_LMP',
        'version': 1,
        'interval_minutes': 60,
        'value_column': 'MW',
        'has_greenhouse_gas': True,
        'report_name': 'Locational Marginal Prices'
    },
    'RTM': {
        'queryname': 'PRC_INTVL_LMP',
        'version': 3,
        'interval_minutes': 5,
        'value_column': 'VALUE',
        'has_greenhouse_gas': True,
        'report_name': 'Interval Locational Marginal Prices'
    },
    'HASP': {
        'queryname': 'PRC_HASP_LMP',
        'version': 1,
        'interval_minutes': 15,
        'value_column': 'MW',
        'has_greenhouse_gas': False,
        'report_name': 'Hour Ahead Locational Marginal Prices'
    },
    'FMM': {
        'queryname': 'PRC_RTPD_LMP',
        'api_market_run_id': 'RTPD',  # CAISO API uses RTPD, not FMM
        'version': 2,
        'interval_minutes': 15,
        'value_column': 'PRC',
        'has_greenhouse_gas': True,
        'report_name': 'FMM Locational Marginal Prices'
    }
}

def get_market_config(market_run_id):
    '''
    This method calls the MARKET_CONFIG map and links it to the
     appropriate variables based on market_run_id
    '''
    return MARKET_CONFIG.get(market_run_id, MARKET_CONFIG['DAM'])

# =============================================================================
# DATA PROCESSING FUNCTION
# =============================================================================
def backend(market_run_id, startdate, enddate, include_ghg=True):
    '''
    This method runs everything the GUI does after the submit button
    is pressed. It calls the API, pulls the cleaned data into an
    excel file, and adds 3 analysis sheets: hourly average, monthly
    average, and summary statistics. When include_ghg is False, the
    Greenhouse Gas component is subtracted out of LMP (and the columns
    are relabeled) rather than counted toward it.
    '''
    node = node_var.get()
    node = node.replace(' ', '')
    # Nodes in the order the user typed them (used for numbering per-node tabs/charts)
    node_list = [n for n in node.split(',') if n]

    # API Calling Function: returns a raw DataFrame for the given date range, or None on error
    def pull_request(chunk_start, chunk_end):
        start_int = int(chunk_start.strftime('%Y%m%d'))
        end_int = int(chunk_end.strftime('%Y%m%d'))

        url = "http://oasis.caiso.com/oasisapi/SingleZip"
        params = {
            "resultformat": 6,  # Resultformat should always be 6- it creates a CSV.
            "queryname": queryname,  # Represents the report name, based on market_run_id
            "startdatetime": f'{start_int}T07:00-0000',  # Dates start at 7 hour for MST purposes
            "enddatetime": f'{end_int}T07:00-0000',
            "market_run_id": api_market_run_id,
            "version": version,
            "node": node,
        }

        # --- DEBUG: uncomment the lines below to print API details to the terminal ---
        # print(f'\n--- API REQUEST ---')
        # print(f'URL: {url}')
        # print(f'Params: {params}')
        try:
            response = requests.get(url, params=params, timeout=1000)
            # print(f'HTTP status: {response.status_code}')
            # print(f'Content-Type: {response.headers.get("Content-Type")}')
            # print(f'Content length: {len(response.content)} bytes')
            with ZipFile(BytesIO(response.content)) as z:
                # print(f'Zip contents: {z.namelist()}')
                for filename in z.namelist():
                    with z.open(filename) as f:
                        df = pd.read_csv(f)
                        # print(f'CSV columns: {list(df.columns)}')
                        # print(f'CSV shape: {df.shape}')
                        if '<?xml version="1.0" encoding="UTF-8"?>' in df.columns:
                            # print('Result: XML error response from API')
                            update_status('No data returned. Check node name(s), date range, and market type.')
                            return None
                        # print('Result: valid data received')
                        return df
            # print('Result: zip had no files')
        except ValidationError as e:
            update_status(f'Param error: {e}')
        except Exception as e:
            # import traceback; traceback.print_exc()
            update_status(f'Request error: {type(e).__name__}: {e}')
        return None

    # Cleans a raw chunk DataFrame in memory: drops columns, converts UTC→MST, splits date fields
    def clean_df(df):
        all_drop = ['NODE_ID_XML', 'NODE_ID', 'PNODE_RESMRID',  # Dropping unneeded columns
                    'OPR_DT', 'OPR_HR', 'OPR_INTERVAL',
                    'XML_DATA_ITEM', 'POS', 'GROUP',
                    'GRP_TYPE', 'MARKET_RUN_ID', 'Unnamed: 0',
                    'INTERVAL_START_TIME']
        valid_columns = [col for col in all_drop if col in df.columns]
        df = df.drop(columns=valid_columns)  # Dropping conditionally
        df = df.sort_values(['INTERVALSTARTTIME_GMT'])

        # Changing the timezone to MST
        df['INTERVALSTARTTIME_GMT'] = df['INTERVALSTARTTIME_GMT'].str.replace('-00:00', '')
        df['INTERVALSTARTTIME_GMT'] = df['INTERVALSTARTTIME_GMT'].str.replace('T', ' ')
        df['INTERVALENDTIME_GMT'] = df['INTERVALENDTIME_GMT'].str.replace('-00:00','')
        df['INTERVALENDTIME_GMT'] = df['INTERVALENDTIME_GMT'].str.replace('T',' ')
        # Shifting the hours back by 7 to align with MST
        df['INTERVALSTARTTIME_GMT'] = pd.to_datetime(df['INTERVALSTARTTIME_GMT'])
        df['INTERVALSTARTTIME_GMT'] = df['INTERVALSTARTTIME_GMT'] - timedelta(hours=7)
        df['INTERVALENDTIME_GMT'] = pd.to_datetime(df['INTERVALENDTIME_GMT'])
        df['INTERVALENDTIME_GMT'] = df['INTERVALENDTIME_GMT'] - timedelta(hours=7)
        df.rename(columns={'INTERVALSTARTTIME_GMT': 'INTERVALSTARTTIME_MST',
                           'INTERVALENDTIME_GMT': 'INTERVALENDTIME_MST'}, inplace=True)

        # Conditional cleaning based on column name for price
        if 'LMP_TYPE' in df.columns:
            df['LMP_TYPE'] = df['LMP_TYPE'].replace({'LMP':'LMP', 'MCC':'Congestion',
                                                     'MCE':'Energy', 'MCL':'Loss',
                                                     'MGHG':'Greenhouse Gas'})

        # Splitting date into smaller columns for readability and grouping
        df['INTERVALSTARTTIME_MST'] = df['INTERVALSTARTTIME_MST'].astype(str)
        df[['Year', 'Month', 'Day']] = df['INTERVALSTARTTIME_MST'].str.split('-',expand=True)
        df[['Day', 'Time']] = df['Day'].str.split(' ',expand=True)
        df[['Hour (MST)', 'Minute', 'Seconds']] = df['Time'].str.split(':',expand=True)
        df = df.drop(columns=['Time', 'Seconds'])
        return df

    # Saves the current matplotlib figure to an in-memory PNG buffer (no temp files)
    def fig_to_buf():
        '''
        This method renders the current matplotlib figure to a BytesIO
        buffer so it can be embedded directly, avoiding temp .png files.
        '''
        buf = BytesIO()
        plt.savefig(buf, format='png')
        plt.close()  # Closing plt so it doesn't combine with the next chart
        buf.seek(0)
        return buf

    # Finding interval rows that are missing from the DF and adding in the missing values
    def build_filled_report(df, market_run_id):
        '''
        This method finds all of the missing intervals not found in the
        df. It fills in the missing values and backfills LMP values.
        Works entirely in memory and returns the filled report DataFrame.
        '''
        config = get_market_config(market_run_id)
        interval_minutes = config['interval_minutes']

        df = df.copy()
        # Finding and creating rows for missing intervals
        dt = pd.to_datetime(df['INTERVALSTARTTIME_MST'])
        df['INTERVALSTARTTIME_MST'] = dt.dt.floor(f'{interval_minutes}min')
        # Finding all intervals I should have
        full_range = pd.date_range(start=df['INTERVALSTARTTIME_MST'].min(),
                                   end=df['INTERVALSTARTTIME_MST'].max(),
                                   freq=f'{interval_minutes}min')
        full_df = pd.DataFrame({'INTERVALSTARTTIME_MST': full_range})  # New df for all intervals
        if 'Greenhouse Gas' in df.columns:  # Conditional logic for HASP (no greenhouse gas)
            result = full_df.merge(
                df[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST',
                    'NODE', 'Year', 'Month', 'Hour (MST)', 'Minute',
                    'Congestion', 'Energy', 'Loss', 'Greenhouse Gas',
                    'LMP']],
                on='INTERVALSTARTTIME_MST',
                how='outer'
            )
        else:
            result = full_df.merge(
                df[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST',
                    'NODE', 'Year', 'Month', 'Hour (MST)', 'Minute',
                    'Congestion', 'Energy', 'Loss', 'LMP']],
                on='INTERVALSTARTTIME_MST',
                how='outer'
            )
        # Filling in missing values
        interval_end = result['INTERVALSTARTTIME_MST'] + pd.Timedelta(minutes=interval_minutes)
        result['INTERVALENDTIME_MST'] = result['INTERVALENDTIME_MST'].fillna(interval_end)
        result = result.sort_values(['INTERVALSTARTTIME_MST', 'NODE'])  # Sort by node to backfill
        result['NODE'] = result['NODE'].bfill()
        result['INTERVALSTARTTIME_MST'] = result['INTERVALSTARTTIME_MST'].astype(str)
        if result['INTERVALENDTIME_MST'].isnull:  # If it has missing values, fill them in!
            result[['Year', 'Month', 'Day']] = result['INTERVALSTARTTIME_MST'].str.split('-',
                expand=True)
            result[['Day', 'Time']] = result['Day'].str.split(' ',expand=True)
            result[['Hour (MST)', 'Minute', 'Seconds']] = result['Time'].str.split(':',expand=True)
        result = result.drop(columns=['Time', 'Seconds']).sort_values(['Hour (MST)', 'Minute'])

        # Filling in LMP values, backfilling with previous day's same hour and minute intvl price.
        if 'Greenhouse Gas' in result.columns:
            LMP_list = ['Congestion', 'Energy', 'Loss', 'Greenhouse Gas', 'LMP']
        else:
            LMP_list = ['Congestion', 'Energy', 'Loss', 'LMP']
        for column in LMP_list:
            result[column] = result[column].bfill()
        result = result.sort_values('INTERVALSTARTTIME_MST')  # Setting to original sorting
        # Reordering columns
        if 'Greenhouse Gas' in result.columns:
            result = result[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST',
                             'NODE', 'Year', 'Month', 'Day', 'Hour (MST)',
                             'Minute', 'LMP','Congestion', 'Energy', 'Loss',
                             'Greenhouse Gas']]
        else:
            result = result[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST',
                             'NODE', 'Year', 'Month', 'Day', 'Hour (MST)',
                             'Minute', 'LMP','Congestion', 'Energy', 'Loss']]
        result = result.reset_index(drop=True)
        return result

    # Creating the monthly average sheet and chart(s)
    def compute_monthly(df, ordered_nodes):
        '''
        This method computes the (combined) monthly average table and
        builds the monthly line chart(s). With multiple nodes it returns
        a group chart (all node lines overlaid) followed by one chart per
        node. Returns (DataFrame, list of image buffers in stacking order).
        '''
        df = df.copy()
        # Create a new column containing month and year and have this be the X axis
        df['Date'] = df['Month'].astype(str).str.zfill(2) + '/01/' + df['Year'].astype(str)
        df['Date'] = pd.to_datetime(df['Date'], format='%m/%d/%Y')
        df['Date'] = df['Date'].dt.strftime('%m %Y')  # Prints date as month + year (05 2024)
        df = df.sort_values(['Year', 'Month']) # sorts first by year, then by month

        if 'Greenhouse Gas' in df.columns:
            df_avg = df.groupby(['NODE', 'Year', 'Month', 'Date'],
                                as_index=False)[['Congestion', 'Energy',
                                                 'Greenhouse Gas', 'LMP', 'Loss']].mean()
            df_avg = df_avg[['NODE', 'Year', 'Month', 'Date', 'LMP',
                             'Congestion', 'Energy', 'Greenhouse Gas', 'Loss']]
        else:
            df_avg = df.groupby(['NODE', 'Year', 'Month', 'Date'],
                                as_index=False)[['Congestion', 'Energy',
                                                 'LMP', 'Loss']].mean()
            df_avg = df_avg[['NODE', 'Year', 'Month', 'Date', 'LMP',
                             'Congestion', 'Energy', 'Loss']]

        # Draws a single node's monthly line, or all nodes overlaid for the group chart
        def draw_monthly(title, overlay=False):
            plt.figure(figsize=(10,8))
            if overlay:
                for n in ordered_nodes:
                    node_df = df_avg[df_avg['NODE'] == n]
                    plt.plot(node_df['Date'], node_df['LMP'], label=n)
                plt.legend()
            else:
                node_df = df_avg[df_avg['NODE'] == title_node]
                plt.plot(node_df['Date'], node_df['LMP'])
            plt.tick_params(axis='x', labelrotation=45)
            plt.grid()
            plt.title(title)
            plt.ylabel('Avg $/MWh')
            plt.xlabel('Month')
            plt.tight_layout()
            return fig_to_buf()

        imgs = []
        if len(ordered_nodes) > 1:
            imgs.append(draw_monthly('Monthly Average LMP - All Nodes', overlay=True))
            for n in ordered_nodes:
                title_node = n
                imgs.append(draw_monthly(f'Monthly Average LMP {n}'))
        else:
            title_node = ordered_nodes[0]
            imgs.append(draw_monthly(f'Monthly Average LMP {title_node}'))
        return df_avg, imgs

    # Creating hourly average sheet and heatmap(s)
    def compute_hourly(df, ordered_nodes):
        '''
        This method takes the (combined) hourly average of the main report
        and builds the 12x24 heatmap(s). With multiple nodes it returns a
        group heatmap (all nodes averaged) followed by one per node.
        Returns (DataFrame, group below-zero count, {node: count}, list of
        image buffers in stacking order).
        '''
        df = df.copy()
        if 'Greenhouse Gas' in df.columns:
            df_avg = df.groupby(['NODE', 'Year', 'Month', 'Day', 'Hour (MST)'],
                                as_index=False)[['Congestion', 'Energy', 'Greenhouse Gas', 'LMP',
                                                 'Loss']].mean()
            df_avg = df_avg[['NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Congestion', 'Energy',
                             'Greenhouse Gas', 'Loss', 'LMP']]  # Reordering column names
        else:
            df_avg = df.groupby(['NODE', 'Year', 'Month', 'Day', 'Hour (MST)'],
                                as_index=False)[['Congestion', 'Energy', 'LMP', 'Loss']].mean()
            df_avg = df_avg[['NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Congestion', 'Energy',
                             'Loss', 'LMP']]

        # Counting how many hours LMP is below 0, for the group and for each node
        count = (df_avg['LMP'] < 0).sum()
        per_node_counts = {n: int((df_avg.loc[df_avg['NODE'] == n, 'LMP'] < 0).sum())
                           for n in ordered_nodes}

        # Draws a 12x24 heatmap for the given (possibly node-filtered) data
        def draw_heatmap(sub_df, title):
            pivot = sub_df.pivot_table(index='Hour (MST)', columns='Month',
                                       values='LMP', aggfunc='mean')
            plt.figure(figsize=(10,8))
            sns.heatmap(pivot, annot=True, cmap='RdYlGn_r', fmt='.0f', cbar_kws={'label':'$/MWH'})
            plt.title(title)
            plt.tight_layout()
            return fig_to_buf()

        imgs = []
        if len(ordered_nodes) > 1:
            imgs.append(draw_heatmap(df, '12x24 Heatmap - All Nodes'))
            for n in ordered_nodes:
                imgs.append(draw_heatmap(df[df['NODE'] == n], f'12x24 Heatmap {n}'))
        else:
            imgs.append(draw_heatmap(df, f'12x24 Heatmap {ordered_nodes[0]}'))
        return df_avg, count, per_node_counts, imgs

    # Creating summary statistics table
    def compute_summary(df):
        '''
        This method computes summary stats for the four LMP values and
        returns the describe() DataFrame.
        '''
        if 'Greenhouse Gas' in df.columns:
            cols = ['Congestion', 'Energy', 'Loss', 'Greenhouse Gas', 'LMP']
        else:
            cols = ['Congestion', 'Energy', 'Loss', 'LMP']
        desc = df[cols].describe()  # Finding what I want to display
        desc = desc.round(4)
        return desc

    # Creating duration charts and the hidden data they plot from
    def compute_duration(df, label):
        '''
        This method creates a duration chart for the entire report, as
        well as two zoomed charts, for the first and last 5%. Returns
        (hidden-data DataFrame, list of (image buffer, cell) tuples).
        '''
        # Cleaning to get chart columns
        df = df.copy().sort_values('LMP', ascending=False)
        duration_counts = df['LMP'].value_counts()
        total_count = df['LMP'].value_counts().sum()
        df['duration_count'] = df['LMP'].map(duration_counts)  # Represents count of that value
        df['percent'] = df['duration_count']/total_count  # % of how often value appears in the df
        chart_lmp = df[['LMP', 'percent']].drop_duplicates().copy()
        chart_lmp['xval'] = chart_lmp['percent'].cumsum()
        xval_map = dict(zip(chart_lmp['LMP'], chart_lmp['xval']))
        df['xval'] = df['LMP'].map(xval_map)
        df = df[['LMP', 'xval']]

        duration_imgs = []  # (image buffer, target cell) tuples, embedded during the write phase

        # Creating duration chart
        plt.figure()
        plt.scatter(df['xval'], df['LMP'], s=3)
        plt.axhline(y=0, color ='black')  # Creating 0 axis line
        plt.title(f'Duration Chart {label}')
        plt.ylabel('$/MWh')
        plt.xlabel('% of Time')
        plt.grid()
        duration_imgs.append((fig_to_buf(), 'B19'))

        # Creating lowest 5% zoom
        last5 = df.tail(int(len(df)*.05))
        plt.figure()
        plt.scatter(last5['xval'], last5['LMP'], s=3)
        plt.axhline(y=0, color='black')
        plt.title(f'Lowest 5% Zoom {label}')
        plt.ylabel('$/MWh')
        plt.xlabel('% of Time')
        plt.grid()
        duration_imgs.append((fig_to_buf(), 'L46'))

        # Creating highest 5% zoom
        first5 = df.head(int(len(df)*.05))
        plt.figure()
        plt.scatter(first5['xval'], first5['LMP'], s=3)
        plt.axhline(y=0, color='black')
        plt.title(f'Highest 5% Zoom {label}')
        plt.ylabel('$/MWh')
        plt.xlabel('% of Time')
        plt.grid()
        duration_imgs.append((fig_to_buf(), 'B46'))

        return df, duration_imgs

    # Formats all numbers in the given column range to have two decimal places
    def format_number_cells(ws, min_col, max_col):
        '''
        This method formats the numerical values to two decimal places
        on an already-open worksheet object.
        '''
        for row in ws.iter_rows(min_col=min_col, max_col=max_col, min_row=2):
            for cell in row:
                cell.number_format = '0.00'

    # Returning the appropriate LMP columns based on market_run_id
    def get_lmp_columns(has_greenhouse_gas=True):
        '''
        This method creates column lists based on whether
        'Greenhouse Gas' is a column.
        '''
        base_columns = ['Congestion', 'Energy', 'Loss', 'LMP']
        if has_greenhouse_gas:
            base_columns.insert(-1, 'Greenhouse Gas')  # Inserting before LMP
        return base_columns

    # Filtering columns depending on if it has greenhouse gas or not
    def get_ordered_columns(has_greenhouse_gas=True):
        base_columns = ['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST',
                        'NODE', 'Year', 'Month', 'Day', 'Hour (MST)',
                        'Minute']
        lmp_columns = get_lmp_columns(has_greenhouse_gas)
        return base_columns + lmp_columns


    # Preparing variables from user input
    config = get_market_config(market_run_id)
    queryname = config['queryname']
    version = config['version']
    api_market_run_id = config.get('api_market_run_id', market_run_id)
    startdate = datetime.strptime(startdate, '%m/%d/%y').date()
    enddate = datetime.strptime(enddate, '%m/%d/%y').date()
    # Filename-safe date labels (captured before enddate +1 and before the loop mutates startdate)
    start_label = startdate.strftime('%m-%d-%Y')
    end_label = enddate.strftime('%m-%d-%Y')
    enddate = enddate + timedelta(days=1)  # Adding on a day to pull so it gets the full last day
    difference = enddate - startdate
    days = difference.days  # Making a counter for my loop bc .days is readonly

    # Download data in 30-day chunks, keeping all DataFrames in memory (no temp files)
    df_list = []
    timestamp = datetime.now().strftime('%m-%d-%Y %H%M')
    total_chunks = math.ceil(days / 30)
    chunk_num = 0

    while days > 0:  # Until we reach the end date
        chunk_num += 1
        if days < 30:  # When we get below 30 days left
            update_status(
                f'Downloading chunk {chunk_num} of {total_chunks} ({startdate} → {enddate})...'
            )
            chunk_df = pull_request(startdate, enddate)
            if chunk_df is not None:
                df_list.append(clean_df(chunk_df))
            break  # Ending loop once I get to the end date
        nextdate = startdate + timedelta(days=30)  # Creating a chunk
        update_status(
            f'Downloading chunk {chunk_num} of {total_chunks} ({startdate} → {nextdate})...'
        )
        chunk_df = pull_request(startdate, nextdate)
        if chunk_df is not None:
            df_list.append(clean_df(chunk_df))
        days -= 30
        startdate = nextdate
        if days > 0:  # Skip sleep after the last chunk
            update_status(
                f'Waiting 5s before next request (chunk {chunk_num} of {total_chunks})...'
            )
            time.sleep(5)  # Waiting 10 seconds to avoid query limits

    if not df_list:
        update_status('No data returned. Check node name(s), date range, and market type.')
        root.after(0, lambda: sub_btn.configure(state='normal'))
        return

    update_status('Combining data...')
    # Combining chunks and cleaning them up a little
    df_combined = pd.concat(df_list, ignore_index=True)
    cond_drop = ['Unnamed: 0.1', 'Unnamed: 0']
    conditional_drop = [col for col in cond_drop if col in df_combined.columns]
    df_combined = df_combined.drop(columns=conditional_drop)
    df_combined = df_combined.drop_duplicates()  # Dropping duplicates

    # Pivoting table and reordering columns (for first sheet)
    config = get_market_config(market_run_id)
    value_column = config['value_column']
    has_greenhouse_gas = config['has_greenhouse_gas']

    # Breaking out LMP_TYPE columns, keeping the other indexed columns
    if value_column in df_combined.columns:
        df_combined = pd.pivot_table(df_combined, values=value_column,
                                     index=['INTERVALSTARTTIME_MST',
                                            'INTERVALENDTIME_MST',
                                            'NODE', 'Year', 'Month',
                                            'Day', 'Hour (MST)',
                                            'Minute'], columns='LMP_TYPE').reset_index()
        df_combined = df_combined[get_ordered_columns(has_greenhouse_gas)]

    # Fill missing intervals in memory (no Excel round-trips)
    update_status('Filling missing intervals...')
    df_report = build_filled_report(df_combined, market_run_id)

    # Greenhouse Gas handling. CAISO reports LMP as the total, i.e.
    # LMP = Energy + Congestion + Loss + GHG. If the user unchecked "Include GHG",
    # subtract the GHG component back out of LMP so every downstream table/chart
    # reflects LMP excluding GHG. The GHG values are kept (never re-pulled) and the
    # columns are relabeled on output so the exclusion is visible.
    exclude_ghg = ('Greenhouse Gas' in df_report.columns) and not include_ghg
    if exclude_ghg:
        df_report['LMP'] = df_report['LMP'] - df_report['Greenhouse Gas']
        out_rename = {'LMP': 'LMP excluding GHG', 'Greenhouse Gas': 'GHG - Not Included'}
    else:
        out_rename = {}

    def out(frame):  # Applies the GHG relabeling to a DataFrame just before writing
        return frame.rename(columns=out_rename) if out_rename else frame

    # Node order as typed, keeping only nodes actually present in the returned data
    report_nodes = list(df_report['NODE'].astype(str).unique())
    ordered_nodes = [n for n in node_list if n in report_nodes]
    ordered_nodes += [n for n in report_nodes if n not in ordered_nodes]  # defensive
    if not ordered_nodes:
        ordered_nodes = report_nodes
    multi_node = len(ordered_nodes) > 1

    # Build every analysis table + chart in memory before touching Excel
    update_status('Generating Monthly Average...')
    df_monthly, monthly_imgs = compute_monthly(df_report, ordered_nodes)

    update_status('Generating Hourly Average...')
    df_hourly, below_zero_count, per_node_counts, heatmap_imgs = compute_hourly(
        df_report, ordered_nodes)

    update_status('Generating Summary Statistics...')
    group_label = 'All Nodes' if multi_node else ordered_nodes[0]
    desc = compute_summary(df_report)

    update_status('Generating Duration Charts...')
    df_duration, duration_imgs = compute_duration(df_report, group_label)

    # Per-node summary pages (only when more than one node — otherwise the group IS the node)
    node_pages = []  # (tab_name, node_name, desc, below-zero count, duration image list)
    if multi_node:
        for i, n in enumerate(ordered_nodes, start=1):
            update_status(f'Generating Summary Statistics for {n}...')
            node_df = df_report[df_report['NODE'].astype(str) == n]
            node_desc = compute_summary(node_df)
            _, node_dur_imgs = compute_duration(node_df, n)
            node_pages.append((f'Summary Statistics - Node {i}', n, node_desc,
                               per_node_counts.get(n, 0), node_dur_imgs))

    # Writes the shared Summary-Statistics page layout onto an already-created sheet
    def decorate_summary_sheet(ws, count, dur_imgs, node_label):
        ws['A1'] = 'Summary Statistics'
        ws['B1'] = node_label  # Node name next to the title
        ws['A14'] = 'Number of hours LMP is below 0:'
        ws['A15'] = int(count)
        ws['A17'] = 'Duration Curve'
        for img_buf, cell_ref in dur_imgs:
            ws.add_image(XLImage(img_buf), cell_ref)
        for coord in ('A1', 'B1', 'A14', 'A17', 'M17'):
            ws[coord].font = Font(bold=True)
        format_number_cells(ws, 2, 6)

    # Stacks a list of chart buffers vertically down one column
    def stack_images(ws, buffers, col, start_row, step=42):
        for idx, buf in enumerate(buffers):
            ws.add_image(XLImage(buf), f'{col}{start_row + idx * step}')

    # Single-pass write: assemble the whole workbook in memory, then save once.
    # This replaces ~20 openpyxl load/save cycles (which re-serialized every
    # embedded chart each time) with one serialization.
    update_status('Writing report to Excel...')
    file = f'{output_file_path}/{market_run_id} LMP {start_label} to {end_label} ({timestamp}).xlsx'
    with pd.ExcelWriter(file, engine='openpyxl') as writer:
        out(df_report).to_excel(writer, sheet_name='Report', index=False)
        out(df_monthly).to_excel(writer, sheet_name='Monthly Average', index=False)
        out(df_hourly).to_excel(writer, sheet_name='Hourly Average', index=False)
        out(desc).to_excel(writer, sheet_name='Summary Statistics', startrow=1,
                           header=True, index=True)
        for tab_name, _, node_desc, _, _ in node_pages:
            out(node_desc).to_excel(writer, sheet_name=tab_name, startrow=1,
                                    header=True, index=True)
        out(df_duration).to_excel(writer, sheet_name='Hidden Duration Chart Data', index=False)

        ws_report = writer.sheets['Report']
        ws_monthly = writer.sheets['Monthly Average']
        ws_hourly = writer.sheets['Hourly Average']
        ws_summary = writer.sheets['Summary Statistics']
        ws_duration = writer.sheets['Hidden Duration Chart Data']

        # Group Summary Statistics page
        decorate_summary_sheet(ws_summary, below_zero_count, duration_imgs, group_label)

        # Per-node Summary Statistics pages
        for tab_name, node_name, _, node_count, node_dur_imgs in node_pages:
            decorate_summary_sheet(writer.sheets[tab_name], node_count,
                                   node_dur_imgs, node_name)

        # Embedding the monthly/hourly charts, stacked (group first, then per node)
        stack_images(ws_monthly, monthly_imgs, 'J', 1)
        stack_images(ws_hourly, heatmap_imgs, 'L', 3)

        # Hiding the raw duration-curve data sheet
        ws_duration.sheet_state = 'hidden'

        # Number formatting (two decimals) on the price columns. Ranges cover the
        # optional Greenhouse Gas column; on markets without it the extra column is
        # empty and formatting it is harmless.
        format_number_cells(ws_monthly, 4, 10)
        format_number_cells(ws_hourly, 6, 10)
        format_number_cells(ws_report, 9, 13)

    root.after(0, lambda: (
        status_lbl.configure(text='Finished!'),
        sub_btn.configure(state='normal')
    ))


# =============================================================================
# GUI WIDGET FUNCTIONS
# =============================================================================
def update_status(text):
    root.after(0, lambda: status_lbl.configure(text=text))

def submit():
    '''
    After user gives all inputs, this method runs all of the backend code.
    '''
    sub_btn.configure(state='disabled')
    update_status('Starting...')
    market_run_id = MRIDDropdown.get()  # Grabbing market_run_id based on user input
    include_ghg = ghg_var.get()  # Whether GHG is counted toward LMP
    t = threading.Thread(
        target=lambda: backend(market_run_id, startdate, enddate, include_ghg),
        daemon=True
    )
    t.start()

def find_start_date():
    '''
    Creates start date based on user input
    '''
    global startdate
    startdate = cal.get_date()
    startdate_label.configure(text=f'Start date: {startdate}')

def find_end_date():
    '''
    Creates start date based on user input
    '''
    global enddate
    enddate = cal.get_date()
    enddate_label.configure(text=f'End date: {enddate}')

def select_output_file():
    '''
    Allows user to select an output location and creates a path
    '''
    global output_file_path
    directory = filedialog.askdirectory(title='Select output directory')
    if directory:
        output_file_path = directory
        output_file_label.configure(text=directory)  # Displaying file path
    else:
        output_file_label.configure(text='No directory selected yet')  # If submitted w/o filepath

def update_report_lbl(choice):
    '''
    Displays the report name based on the user chosen market type
    '''
    config = get_market_config(choice)
    report_name = config.get('report_name', 'Unknown Report')

    report_lbl.configure(text=f'{report_name}')
    # Disable the GHG checkbox for markets without a GHG component (e.g., HASP)
    if config.get('has_greenhouse_gas', False):
        ghg_checkbox.configure(state='normal')
    else:
        ghg_checkbox.configure(state='disabled')
    root.update()


# =============================================================================
# TKINTER PROGRAM
# =============================================================================
root = CTk()  # Initializing window
root.geometry('800x600')
set_appearance_mode('light')

node_var = tk.StringVar()
ghg_var = tk.BooleanVar(value=True)  # Include Greenhouse Gas in LMP (default: yes)

startdate = None  # Initializing
enddate = None

# =============================================================================
# TKINTER WIDGETS
# =============================================================================
MRID_label = CTkLabel(root, text='Market Type:', font=('Arial',15), text_color='#04033A')
MRIDDropdown = CTkComboBox(master=root, values=['DAM', 'HASP','RTM', 'FMM'],
                           command=update_report_lbl)

report_lbl = CTkLabel(root, text='Locational Marginal Prices', font=('Arial', 15),
                      text_color='#04033A')

cal = Calendar(root, selectmode ='day',
            year=2024, month=1,  # Default setting
            day=1, font=('Arial', 15))

chooseStartDate = CTkButton(root, text='Choose Start Date', command=find_start_date,
                            corner_radius=26, fg_color='#162157', hover_color='#6D7DCF')
chooseEndDate = CTkButton(root, text='Choose End Date', command=find_end_date, corner_radius=26,
                          fg_color='#162157', hover_color='#6D7DCF')

startdate_label = CTkLabel(root, text= 'Start Date: ', font=('Arial', 15), text_color='#04033A')
enddate_label = CTkLabel(root, text='End Date: ', font=('Arial', 15), text_color='#04033A')

node_label = CTkLabel(root, text='Node(s):', font=('Arial', 15), text_color='#04033A')
node_entry = CTkEntry(root, textvariable = node_var, font=('Arial', 15), text_color='#04033A', width=300)

ghg_checkbox = CTkCheckBox(root, text='Include Greenhouse Gas in LMP', variable=ghg_var,
                           onvalue=True, offvalue=False, font=('Arial', 13),
                           text_color='#04033A', fg_color='#162157', hover_color='#6D7DCF')

sub_btn=CTkButton(master=root,text = 'Submit', command=submit, corner_radius=32,
                  fg_color='#162157', hover_color='#6D7DCF')

output_file_button = CTkButton(root, text='Select Output File Path', command=select_output_file,
                               corner_radius=32, fg_color='#162157', hover_color='#6D7DCF')
output_file_label = CTkLabel(root, text='No path selected', font=('Arial', 10),
                             text_color='#04033A', wraplength=220)

status_lbl = CTkLabel(root, text='', font=('Arial', 15), text_color='#04033A')

title_lbl = CTkLabel(root, text='CAISO OASIS DATA', font=('Arial', 20, 'bold'),
                     text_color='#04033A')

# =============================================================================
# TKINTER GRID AND APP START
# =============================================================================
cal.grid(row=6, column=0)
chooseStartDate.grid(row=4, column=0)
chooseEndDate.grid(row=5, column=0)
MRID_label.grid(row=1, column=1)
MRIDDropdown.grid(row=1, column=2)
node_label.grid(row=3, column=1)
node_entry.grid(row=3, column=2)
ghg_checkbox.grid(row=4, column=2)
sub_btn.grid(row=6, column=2)
startdate_label.grid(row=4, column=1)
enddate_label.grid(row=5, column=1)
output_file_button.grid(row=1, column=0)
output_file_label.grid(row=2, column=0)
title_lbl.grid(row=0, column=2)
report_lbl.grid(row=2, column=2)
status_lbl.grid(row=7, column=2)

root.mainloop()  # Performing an infinite loop for the window to display
